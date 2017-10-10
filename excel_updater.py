import openpyxl, csv, time, shutil, os, win32com.client

wb = openpyxl.Workbook()

Attsheet = wb.create_sheet(index=0, title='Att')
Att = open('Attending.csv')
Attcsv = csv.reader(Att)
for row in Attcsv:
    Attsheet.append(row)

PAsheet = wb.create_sheet(index=1, title='PA')
PA = open('PA.csv')
PAcsv = csv.reader(PA)
for row in PAcsv:
    PAsheet.append(row)

Ressheet = wb.create_sheet(index=2, title='Res')
Res = open('Resident.csv')
Rescsv = csv.reader(Res)
for row in Rescsv:
    Ressheet.append(row)

wb.save('Provider formula book.xlsx')
wb.close()

time.sleep(1)

SourcePathName = 'C:/scheduleproj'
FileName = 'provider formula book (leadership).xlsx'

Application = win32com.client.Dispatch("Excel.Application")
Application.Visible = 1
Workbook = Application.Workbooks.open(SourcePathName + '/' + FileName)
Workbook.RefreshAll()
Workbook.Save()
Application.Quit()

time.sleep(1)

wb2 = openpyxl.load_workbook('provider formula book (leadership).xlsx', data_only=True)
Attsheet1 = wb2.get_sheet_by_name('Attcsv')
PAsheet1 = wb2.get_sheet_by_name('PAcsv')
Ressheet1 = wb2.get_sheet_by_name('Rescsv')
Attoversheet = wb2.get_sheet_by_name('AttOverncsv')

outputmain = open('main.csv', 'a', newline='')
with outputmain as f:
    c = csv.writer(f)
    c.writerow(['Provider', 'Att/PA/Res', 'Site', 'Team/Location', 'Start', 'End'])
    for r in Attsheet1.rows:
        c.writerow([cell.value for cell in r])
    for r in PAsheet1.rows:
        c.writerow([cell.value for cell in r])
    for r in Ressheet1.rows:
        c.writerow([cell.value for cell in r])
    for r in Attoversheet.rows:
        c.writerow([cell.value for cell in r])

outputmainNoOn = open('main-no-on.csv', 'a', newline='')
with outputmainNoOn as f:
    c = csv.writer(f)
    c.writerow(['Provider', 'Att/PA/Res', 'Site', 'Team/Location', 'Start', 'End'])
    for r in Attsheet1.rows:
        c.writerow([cell.value for cell in r])
    for r in PAsheet1.rows:
        c.writerow([cell.value for cell in r])
    for r in Ressheet1.rows:
        c.writerow([cell.value for cell in r])
wb2.close()

wb2 = openpyxl.load_workbook('provider formula book (leadership).xlsx', data_only=True)

shtdel1 = wb2.get_sheet_by_name('ATTPARes combined-print')
shtdel2 = wb2.get_sheet_by_name('ATTPARes combined')
wb2.remove_sheet(shtdel1)
wb2.remove_sheet(shtdel2)

wb2.save('provider formula book (leadership)1.xlsx')
wb2.close()

wb2 = openpyxl.load_workbook('provider formula book (leadership)1.xlsx')
combinedPrint = wb2.create_sheet(index=3, title='ATTPARes combined-print')
getMainNoOncsv = open('main-no-on.csv')
readMainNoOn = csv.reader(getMainNoOncsv)
for row in readMainNoOn:
    combinedPrint.append(row)

combined = wb2.create_sheet(index=4, title='ATTPARes combined')
getMainCsv = open('main.csv')
readMain = csv.reader(getMainCsv)
for row in readMain:
    if any(row):
        combined.append(row)

wb2.save('provider formula book (leadership)1.xlsx')
wb2.close()

#copies the final excel file and moves it to main (personal) folder
os.chdir('C:\\')
shutil.copy('C:\\scheduleproj\\provider formula book (leadership)1.xlsx', 'H:\\Personal\\Provider Schedule project\\provider formula book (leadership).xlsx')
